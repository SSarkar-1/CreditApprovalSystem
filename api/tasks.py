from celery import shared_task
import openpyxl
from datetime import datetime
from django.db import connection
from .models import Customer, Loan


def _sync_sequence(table_name, id_column):
    # Point the sequence to the next free id based on current table max.
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT setval(
                pg_get_serial_sequence(%s, %s),
                COALESCE((SELECT MAX({id_column}) FROM {table_name}), 0) + 1,
                false
            )
            """,
            [table_name, id_column],
        )


@shared_task
def ingest_customer_data():
    wb = openpyxl.load_workbook('data/customer_data.xlsx')
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        customer_id, first_name, last_name, age, phone_number, monthly_salary, approved_limit = row
        # Ignore blank or malformed rows from the sheet.
        if (
            customer_id is None
            or first_name is None
            or last_name is None
            or phone_number is None
            or monthly_salary is None
            or approved_limit is None
        ):
            continue

        Customer.objects.update_or_create(
            customer_id=customer_id,
            defaults={
                'first_name': first_name,
                'last_name': last_name,
                'age': age,
                'phone_number': phone_number,
                'monthly_salary': monthly_salary,
                'approved_limit': approved_limit,
            }
        )

    _sync_sequence("api_customer", "customer_id")
    print("Customer data ingested successfully")


@shared_task
def ingest_loan_data():
    wb = openpyxl.load_workbook('data/loan_data.xlsx')
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        customer_id, loan_id, loan_amount, tenure, interest_rate, monthly_repayment, emis_paid_on_time, start_date, end_date = row

        try:
            customer = Customer.objects.get(customer_id=customer_id)
        except Customer.DoesNotExist:
            print(f"Customer {customer_id} not found, skipping loan {loan_id}")
            continue

        if isinstance(start_date, str):
            start_date = datetime.strptime(start_date, '%d-%m-%Y').date()
        if isinstance(end_date, str):
            end_date = datetime.strptime(end_date, '%d-%m-%Y').date()

        Loan.objects.update_or_create(
            loan_id=loan_id,
            defaults={
                'customer': customer,
                'loan_amount': loan_amount,
                'tenure': tenure,
                'interest_rate': interest_rate,
                'monthly_repayment': monthly_repayment,
                'emis_paid_on_time': emis_paid_on_time,
                'start_date': start_date,
                'end_date': end_date,
            }
        )

    _sync_sequence("api_loan", "loan_id")
    print("Loan data ingested successfully")
